# -*- coding: utf-8 -*-
"""최신 마스터 엑셀 → master_pool.json 추출."""
from __future__ import annotations

import json
import math
import re
from pathlib import Path

ROOT = Path(__file__).resolve().parent
DATA = ROOT / "data"
MASTER_XLSX = DATA / "응급의학과_봉직공고_최종마스터.xlsx"
OUT_JSON = DATA / "master_pool.json"

SHEET_MAIN = "1.마스터_월수령순"
SHEET_AXES = "16.모델v5.1_중증도·백업불신"

_norm = lambda s: re.sub(r"[\s\-–—·]", "", str(s))


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        return float(v)
    s = str(v).strip().replace(",", "")
    if not s or s in ("미확정", "산출불가", "신설"):
        return None
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    # '잠정2500' 같은 잠정 표기는 결측으로 둔다 (소방)
    if "잠정" in s or "미확" in s:
        return None
    return float(m.group())


def _load_xlsx(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_MAIN not in wb.sheetnames or SHEET_AXES not in wb.sheetnames:
        raise KeyError(f"필수 시트 없음: {wb.sheetnames}")
    main, axes = {}, {}
    only_notes = dict(skipped=[], dup_main=[], dup_axes=[])

    for i, r in enumerate(wb[SHEET_MAIN].iter_rows(values_only=True), 1):
        if r[0] is None or r[1] is None:
            continue
        try:
            int(r[0])
        except (TypeError, ValueError):
            continue
        name = str(r[1]).strip()
        row = dict(
            rank_cash=_num(r[0]),
            h=name,
            region=None if r[2] is None else str(r[2]),
            zone=None if r[3] is None else str(r[3]),
            er=None if r[4] is None else str(r[4]),
            em=_num(r[5]),
            tot=_num(r[6]),
            pp=_num(r[7]),
            hrs=_num(r[8]),
            net_base_eok=_num(r[9]),
            net_inc_eok=_num(r[10]),
            cash=_num(r[11]),
            unit_raw=_num(r[12]),
            backup=None if r[13] is None else str(r[13]),
            legal=None if r[14] is None else str(r[14]),
            note=None if len(r) < 16 or r[15] is None else str(r[15]),
            src_row_main=i,
        )
        k = _norm(name)
        if k in main:
            only_notes["dup_main"].append(name)
        main[k] = row

    for i, r in enumerate(wb[SHEET_AXES].iter_rows(values_only=True), 1):
        if r[0] is None or r[1] is None:
            continue
        try:
            int(r[0])
        except (TypeError, ValueError):
            continue
        name = str(r[1]).strip()
        row = dict(
            h_orig=name,
            region_axes=None if r[2] is None else str(r[2]),
            grade=None if r[3] is None else str(r[3]),
            acu=_num(r[4]),
            struct=_num(r[5]),
            claim=_num(r[6]),
            mismatch=_num(r[7]),
            safe=_num(r[8]),
            pph=_num(r[9]),
            effhr=_num(r[10]),
            effload=_num(r[11]),
            pay_v51=_num(r[12]),
            sys=_num(r[13]),
            wlb=_num(r[14]),
            acu_val=_num(r[15]) if len(r) > 15 else None,
            src_row_axes=i,
        )
        extra = []
        for j in range(16, min(len(r), 18)):
            extra.append(_num(r[j]) if not isinstance(r[j], str) else r[j])
        if extra:
            row["axes_extra"] = extra
        k = _norm(name)
        if k in axes:
            only_notes["dup_axes"].append(name)
        axes[k] = row

    pool = []
    for k, v in main.items():
        if k not in axes:
            continue
        d = dict(v)
        d.update({kk: vv for kk, vv in axes[k].items() if kk != "h_orig"})
        d["key"] = k
        pool.append(d)

    meta = dict(
        source=Path(path).name,
        as_of="2026-08-18",
        n_main=len(main),
        n_axes=len(axes),
        n_merged=len(pool),
        only_main=[main[k]["h"] for k in main if k not in axes],
        only_axes=[axes[k]["h_orig"] for k in axes if k not in main],
        dup_main=only_notes["dup_main"],
        dup_axes=only_notes["dup_axes"],
    )
    wb.close()
    return pool, meta


def extract(path=None, out=None):
    pool, meta = _load_xlsx(path or MASTER_XLSX)
    payload = {"meta": meta, "pool": pool}
    dest = Path(out) if out else OUT_JSON
    dest.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return dest, meta


def main():
    dest, meta = extract()
    print(json.dumps(meta, ensure_ascii=False, indent=2))
    print(f"wrote {dest}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
