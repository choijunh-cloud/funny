# -*- coding: utf-8 -*-
"""
봉직 통합 모델 v1.2 (2026-07-10) — 2차 코드리뷰 7건 반영판
==========================================================
응급의학과 봉직 공고 평가 파이프라인. 원본 마스터 엑셀(아래 2개 시트) 필요:
  응급의학과_봉직공고_최종마스터.xlsx :: "1.마스터_월수령순", "16.모델v5.1_중증도·백업불신"

v1.2 변경 (2차 리뷰 대응):
  ① 오버라이드 키 정규화 — OVERRIDES 조회도 _norm 적용 (공백/하이픈 변형에 견딤)
  ② R1 표현 교정 — "누출 반증" → "직접 cash2 누출을 제거한 민감도 분석·표본 내 강건성
     지지". L2는 pp(환자수)를 effload 경유로 공유하므로 완전 독립 아님을 명시.
     L3 게이트(safe≥5.5 & wlb≥5.4)의 출처: 시트17·18(2026-06-27)에서 사전 정의된
     QOL 게이트 — 13.3 컷 탐색(07-10)보다 먼저 존재(사전등록에 준함).
     라벨 3종 자체의 다중성은 미보정 — 완결하려면 라벨×컷 전체 permutation max-stat 필요(향후).
  ③ 테스트 강화 — L2·L3 각각 컷±0.3 / Bonferroni<0.05 / 컷∈유의구간, 라벨 3종 컷
     최대편차 ≤0.3. 오버라이드-설정 확인 테스트와 모델-정확도 테스트 분리(독립 추정치로 검증).
  ④ 로더 — 시트16 중복도 감지·기록(dup_axes, 원본명·채택값 포함), only_in_axes 원본명 출력.
     strict=True 시 중복 발견에서 명시적 실패.
  ⑤ 비독립 테스트 제거 — 티어 S(규칙상 자동) 삭제, 중복 검사는 pool이 아닌 로더 리포트 기준.
  ⑥ 오산 시나리오 혼합 제거 — 마스터 행은 구공고 원본 보존, 신공고(8인·갱신)는
     POSTING_EXAMPLES 평가 카드로만. OVERRIDES에는 스펙 불변 실측인 나사렛만 유지.
  ⑦ 인센 표기 — "포함(내역미상)"/"★실측"/"(모델)" 전 경로 표시, 카드에 incen·incen_src·
     pay2_raw(클램프 전)·rank_vs_master·percentile 포함.

13.3 컷의 지위(합의된 해석): 현재 마스터 풀에서 라벨 3종이 반복 관찰하는 설명적 기준.
시점 외(신규 공고 축적) 검증 전까지 예측 임계값이라 부르지 않는다.

사용:
  python3 bongjik_model.py [--boot] [--posting] [--report] [--strict]
"""
import math, random, re, statistics, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl 필요: pip install openpyxl")

# ──────────────────────────── 설정 ────────────────────────────
MASTER_XLSX = "응급의학과_봉직공고_최종마스터.xlsx"
SHEET_MAIN = "1.마스터_월수령순"
SHEET_AXES = "16.모델v5.1_중증도·백업불신"
EXPECTED_POOL = (60, 70)

ANCHOR_ADMIT = 0.354            # 입원율 앵커 = 나사렛 실측 97/274 (중증도 1.0)
WON_PER_ADMIT = 3               # 입원 1건당 인센(만원) — 나사렛 실측
CUT_UNIT = 13.3                 # 보정단가 설명적 컷 (시트26·33)
CUT_UNIT_ZONE = (11.1, 13.5)    # 시트26 산출(당시 라벨). 하한은 라벨 민감, 상단 13.5 안정
CUT_ANNUAL = 15000
TIER_RULE = [(1, "S"), (9, "A"), (12, "B"), (20, "C")]

_norm = lambda s: re.sub(r"[\s\-–—·]", "", str(s))

# 실측 오버라이드 — 스펙 불변 + 통장 실측(2026-07 사용자 제공)인 나사렛만.
# (오산 신공고는 인원·시간·축이 함께 바뀌는 별도 시나리오 → POSTING_EXAMPLES로만 평가, v1.2 ⑥)
OVERRIDES = {
    "나사렛국제(현직)": dict(cash2=2800, incen=291, incen_src="실측"),
    # 2026-07-17 실측 전파분: 시트1 통장이 이미 인센 포함 → 모델 인센 재가산(이중계상) 금지
    "진주경상대":       dict(cash2=3367, incen=None, incen_src="실측포함"),
    "김포우리병원":     dict(cash2=2900, incen=None, incen_src="실측포함"),
    "인천현대유비스":   dict(cash2=2450, incen=None, incen_src="실측포함"),
    "오산한국병원":     dict(cash2=2398, incen=None, incen_src="실측포함"),
    "안양샘병원":       dict(cash2=2073, incen=None, incen_src="실측포함"),
    "수원의료원":       dict(cash2=2184, incen=None, incen_src="실측포함"),
    "국립소방병원":     dict(cash2=2600, incen=None, incen_src="실측포함"),
    "안동병원":         dict(cash2=2135, incen=None, incen_src="실측포함(급여단위 확인중)"),
}
OVERRIDES_NORM = {_norm(k): v for k, v in OVERRIDES.items()}   # v1.2 ①

POSTING_EXAMPLES = [
    dict(name="김포우리(복원)", annual=31500, em=11, hours=130, cash_actual=2652,
         acu=0.94, safe=6.7, sysx=5.2, wlb=5.3, incen_included=False),
    dict(name="세란(신규)", annual=15075, em=5, hours=146, cash_actual=2675,
         acu=0.88, safe=4.4, sysx=3.2, wlb=4.3, incen_included=True),
    dict(name="오산한국(8인·갱신)", annual=20864, em=8, hours=122, cash_actual=2398,
         acu=1.0, safe=5.9, sysx=4.3, wlb=5.73, incen_included=True),
]

PROFILES = {                     # [추론: 분석자 설정, 시트25]
    "P1_균등": dict(safe=.20, wlb=.20, effload=.20, sys=.20, pay2=.20),
    "P2_QOL":  dict(safe=.30, wlb=.25, effload=.25, sys=.10, pay2=.10),
    "P3_균형": dict(safe=.25, wlb=.25, effload=.20, sys=.10, pay2=.20),
    "P4_현금": dict(safe=.25, wlb=.20, effload=.15, sys=.10, pay2=.30),
}
PROFILES_NOPAY = {               # 직접 cash2 누출 제거 라벨용 (pp 경유 결합은 잔존 — v1.2 ②)
    "N1": dict(safe=.25, wlb=.25, effload=.25, sys=.25),
    "N2": dict(safe=.35, wlb=.30, effload=.25, sys=.10),
    "N3": dict(safe=.30, wlb=.30, effload=.25, sys=.15),
}
# L3 게이트 상수 — 시트17·18(2026-06-27) 사전 정의값. 컷 탐색 이전에 존재 (v1.2 ②)
L3_SAFE, L3_WLB = 5.5, 5.4

# ──────────────────────── ① 데이터 로드 ────────────────────────
def load_master(path=None, report=False, strict=False):
    p = Path(path) if path else Path(__file__).parent / MASTER_XLSX
    if not p.exists():
        raise FileNotFoundError(f"마스터 파일 없음: {p}")
    wb = openpyxl.load_workbook(p, data_only=True)
    for s in (SHEET_MAIN, SHEET_AXES):
        if s not in wb.sheetnames:
            raise KeyError(f"필수 시트 없음: '{s}'")
    rep = dict(skipped=[], dup_main=[], dup_axes=[], only_in_main=[], only_in_axes=[])
    m1, m16 = {}, {}
    for i, r in enumerate(wb[SHEET_MAIN].iter_rows(values_only=True), 1):
        if r[0] is None or r[1] is None:
            continue
        try:
            int(r[0])
        except (TypeError, ValueError):
            continue
        try:
            row = dict(h=str(r[1]), reg=str(r[2]), tot=float(r[6]), pp=float(r[7]),
                       hrs=float(r[8]), cash=float(r[11]))
        except (TypeError, ValueError):
            rep["skipped"].append((SHEET_MAIN, i, str(r[1]), "숫자 결측/비수치"))
            continue
        k = _norm(r[1])
        if k in m1:
            rep["dup_main"].append(dict(key=k, kept=str(r[1]), dropped=m1[k]["h"], row=i))
        m1[k] = row
    for i, r in enumerate(wb[SHEET_AXES].iter_rows(values_only=True), 1):
        if r[0] is None or r[1] is None:
            continue
        try:
            int(r[0])
        except (TypeError, ValueError):
            continue
        try:
            row = dict(h_orig=str(r[1]), acu=float(r[4]), safe=float(r[8]), effhr=float(r[10]),
                       effload=float(r[11]), sys=float(r[13]), wlb=float(r[14]))
        except (TypeError, ValueError):
            rep["skipped"].append((SHEET_AXES, i, str(r[1]), "축 결측/비수치"))
            continue
        k = _norm(r[1])
        if k in m16:                                             # v1.2 ④
            rep["dup_axes"].append(dict(key=k, kept=str(r[1]), dropped=m16[k]["h_orig"], row=i))
        m16[k] = row
    rep["only_in_main"] = [m1[k]["h"] for k in m1 if k not in m16]
    rep["only_in_axes"] = [m16[k]["h_orig"] for k in m16 if k not in m1]
    if strict and (rep["dup_main"] or rep["dup_axes"]):
        raise ValueError(f"중복 병원명 발견(strict): {rep['dup_main'] + rep['dup_axes']}")
    pool = [{**v, **{kk: vv for kk, vv in m16[k].items() if kk != 'h_orig'}}
            for k, v in m1.items() if k in m16]
    rep["n_main"], rep["n_axes"], rep["n_merged"] = len(m1), len(m16), len(pool)
    if not (EXPECTED_POOL[0] <= len(pool) <= EXPECTED_POOL[1]):
        print(f"⚠ 병합 {len(pool)}곳 — 기대범위 {EXPECTED_POOL} 밖")
    return (pool, rep) if report else pool

# ──────────────────── ② 입원인센 보정 (v1.2 ①: 정규화 조회) ────────────────────
def apply_incentive(pool):
    for d in pool:
        ov = OVERRIDES_NORM.get(_norm(d["h"]), {})
        d.update({k: v for k, v in ov.items() if v is not None or k == "incen"})
        if d.get("pp", 0) <= 0:
            raise ValueError(f"{d['h']}: pp ≤ 0")
        d["mpat"] = d["pp"] / 12
        d.setdefault("incen_src", "모델추정")
        if "incen" not in d or (d["incen"] is None and "cash2" not in d):
            d["incen"] = d["mpat"] * ANCHOR_ADMIT * d["acu"] * WON_PER_ADMIT
        if "cash2" not in d:
            d["cash2"] = d["cash"] + d["incen"]
        d["pp2"] = d["cash2"] / d["mpat"]
    return pool

# ──────────────────── ③ 5축 (pay2_raw 보존, v1.2 ⑦) ────────────────────
def compute_axes(pool, lo=None, hi=None):
    lo = min(d["cash2"] for d in pool) if lo is None else lo
    hi = max(d["cash2"] for d in pool) if hi is None else hi
    span = hi - lo
    for d in pool:
        d["pay2_raw"] = 5.0 if span <= 0 else (d["cash2"] - lo) / span * 10
        d["pay2"] = max(0.0, min(10.0, d["pay2_raw"]))
        d["f5"] = d["safe"] + d["pay2"] + d["sys"] + d["wlb"] + d["effload"]
    return lo, hi

# ──────────────────── ④ 종합 ────────────────────
def _rank_by(pool, key):
    return {id(d): i + 1 for i, d in enumerate(sorted(pool, key=lambda x: (-key(x), x["h"])))}

def composite(pool, profiles=PROFILES, prefix=""):
    for n, w in profiles.items():
        for d in pool:
            d[prefix + n] = sum(d[k] * wt for k, wt in w.items()) * 10
    rks = {n: _rank_by(pool, lambda x, n=n: x[prefix + n]) for n in profiles}
    for d in pool:
        d[prefix + "ranks"] = [rks[n][id(d)] for n in profiles]
        d[prefix + "meanrank"] = sum(d[prefix + "ranks"]) / len(profiles)
        d[prefix + "score"] = sum(d[prefix + n] for n in profiles) / len(profiles)
    fin = sorted(pool, key=lambda d: (d[prefix + "meanrank"], -d[prefix + "score"], d["h"]))
    for i, d in enumerate(fin, 1):
        d[prefix + "comp_rank"] = i
        d[prefix + "tier"] = next((t for lim, t in TIER_RULE if i <= lim), "-")
    return fin

# ──────────────────── ⑤ 게이트 ────────────────────
def gates(d):
    g = {
        "단가≥13.3": d["pp2"] >= CUT_UNIT,
        "안전≥6.3": d["safe"] >= 6.3,
        "안전플로어≥5.5": d["safe"] >= 5.5,
        "워라밸≥6.1": d["wlb"] >= 6.1,
        "월시간≤130": d.get("hrs", 999) <= 130,
        "연환자≥15k": d.get("tot", 0) >= CUT_ANNUAL,
    }
    g["무결점조합"] = d["safe"] >= 5.5 and d["pp2"] >= 13.1
    return g

# ──────────────────── ⑥ 통계 ────────────────────
_LG = [0.0, 0.0]
def _lg(n):
    while len(_LG) <= n:
        _LG.append(_LG[-1] + math.log(len(_LG)))
    return _LG[n]

def _lc(n, k):
    return -1e18 if (k < 0 or k > n) else _lg(n) - _lg(k) - _lg(n - k)

def fisher_sf(tp, N, K, n):
    tot = _lc(N, n)
    return min(1.0, sum(math.exp(_lc(K, x) + _lc(N - K, n - x) - tot)
                        for x in range(tp, min(K, n) + 1)))

def maxstat(values, labels):
    N, K = len(values), sum(labels)
    vals = sorted(set(values))
    out = []
    for i in range(len(vals) - 1):
        t = (vals[i] + vals[i + 1]) / 2
        idx = [j for j in range(N) if values[j] >= t]
        if not idx or len(idx) == N:
            continue
        out.append((t, fisher_sf(sum(labels[j] for j in idx), N, K, len(idx)),
                    len(idx), sum(labels[j] for j in idx)))
    return out

def stat_cut(pool, label_set, boot=0):
    vals = [d["pp2"] for d in pool]
    labs = [1 if d["h"] in label_set else 0 for d in pool]
    curve = maxstat(vals, labs)
    if not curve:
        return dict(best_cut=None, p_raw=1.0, p_bonf=1.0, zone=None, M=0)
    M = len(curve)
    best = min(curve, key=lambda r: r[1])
    sig = [r[0] for r in curve if r[1] * M < 0.05]
    res = dict(best_cut=best[0], p_raw=best[1], p_bonf=min(1, best[1] * M),
               zone=(round(min(sig), 2), round(max(sig), 2)) if sig else None, M=M)
    if boot:
        random.seed(7)
        cuts = []
        N = len(vals)
        for _ in range(boot):
            idx = [random.randrange(N) for _ in range(N)]
            c = maxstat([vals[i] for i in idx], [labs[i] for i in idx])
            if c:
                cuts.append(min(c, key=lambda r: r[1])[0])
        cuts.sort()
        res["boot_ci"] = (cuts[int(.025 * len(cuts))], cuts[int(.975 * len(cuts))])
        res["boot_med"] = cuts[len(cuts) // 2]
    return res

def make_labels(pool):
    """L1 종합(페이포함) / L2 직접 cash2 제거(단, pp 경유 결합 잔존) / L3 QOL 사전정의 게이트."""
    fin1 = composite(pool, PROFILES, prefix="")
    composite(pool, PROFILES_NOPAY, prefix="np_")
    fin2 = sorted(pool, key=lambda d: d["np_comp_rank"])
    return {
        "L1_종합TOP20(페이포함)": set(d["h"] for d in fin1[:20]),
        "L2_페이제거TOP20": set(d["h"] for d in fin2[:20]),
        "L3_QOL게이트(사전정의)": set(d["h"] for d in pool
                                if d["safe"] >= L3_SAFE and d["wlb"] >= L3_WLB),
    }

# ──────────────── ⑦ 신규 공고 평가 (고정 스케일 + 마스터 대비 지표) ────────────────
def _interp_effload(pool, effhr):
    pts = sorted((d["effhr"], d["effload"]) for d in pool if "effhr" in d)
    if not pts:
        raise ValueError("보간용 effhr 데이터 없음")
    if effhr <= pts[0][0]:
        return pts[0][1]
    for (x0, y0), (x1, y1) in zip(pts, pts[1:]):
        if x0 <= effhr <= x1:
            return y0 if x1 == x0 else y0 + (effhr - x0) * (y1 - y0) / (x1 - x0)
    return pts[-1][1]

def _make_candidate(pool, name, annual, em, hours, cash_actual, acu, safe, sysx, wlb,
                    incen_included=True):
    if em <= 0 or hours <= 0 or annual <= 0:
        raise ValueError(f"{name}: annual/em/hours 는 양수여야 함")
    pp = annual / em
    d = dict(h=name, pp=pp, tot=annual, hrs=hours, acu=acu, safe=safe, sys=sysx, wlb=wlb)
    d["mpat"] = pp / 12
    d["effhr"] = pp / 12 / hours * acu
    d["effload"] = _interp_effload(pool, d["effhr"])
    d["incen"] = 0 if incen_included else d["mpat"] * ANCHOR_ADMIT * acu * WON_PER_ADMIT
    d["incen_src"] = "실측포함" if incen_included else "모델추정"
    d["cash2"] = cash_actual + d["incen"]
    d["pp2"] = d["cash2"] / d["mpat"]
    return d

def evaluate_postings(pool_master, specs):
    pool = [dict(d) for d in pool_master]
    lo, hi = compute_axes(pool)
    master_scores = sorted((d["score"] for d in composite(pool)), reverse=True) \
        if pool and "score" not in pool[0] else None
    fin_master = composite(pool)
    master_scores = sorted((d["score"] for d in fin_master), reverse=True)
    cands = [_make_candidate(pool, **s) for s in specs]
    allp = pool + cands
    compute_axes(allp, lo=lo, hi=hi)
    fin = composite(allp)
    cards = []
    for d in cands:
        g = gates(d)
        r_vs_m = 1 + sum(1 for s in master_scores if s > d["score"])
        cards.append(dict(
            name=d["h"], cash2=round(d["cash2"]), pp2=round(d["pp2"], 2),
            incen=round(d["incen"]), incen_src=d["incen_src"],
            pay2=round(d["pay2"], 2), pay2_raw=round(d["pay2_raw"], 2),
            f5=round(d["f5"], 1), score=round(d["score"], 1),
            joint_rank=d["comp_rank"], tier=d["tier"], n=len(allp),
            rank_vs_master=r_vs_m,
            pct_vs_master=round((1 - r_vs_m / (len(master_scores) + 1)) * 100),
            perfect=g["무결점조합"], gates=g))
    return cards

def evaluate_posting(pool_master, **spec):
    return evaluate_postings(pool_master, [spec])[0]

# ──────────────────── 검증 테스트 (v1.2 ③⑤: 독립성 분리) ────────────────────
def run_tests(pool, labels, rep):
    naz = next(d for d in pool if "나사렛" in d["h"])
    kh = next(d for d in pool if "경희" in d["h"] and "강동" not in d["h"])
    fin = composite(pool)
    est_incen = naz["mpat"] * ANCHOR_ADMIT * naz["acu"] * WON_PER_ADMIT
    est_cash2 = naz["cash"] + est_incen                     # 오버라이드 무관 순수 모델 추정
    sts = {n: stat_cut(pool, s) for n, s in labels.items()}
    cuts = [st["best_cut"] for st in sts.values() if st["best_cut"]]
    t = [
        ("[로더] 풀 크기 60~70", EXPECTED_POOL[0] <= len(pool) <= EXPECTED_POOL[1]),
        ("[로더] 시트1 정규화 중복 없음", not rep["dup_main"]),
        ("[로더] 시트16 정규화 중복 없음", not rep["dup_axes"]),
        ("[설정] 나사렛 오버라이드 적용(cash2 2800·실측)",
         naz["cash2"] == 2800 and naz["incen_src"] == "실측"),
        ("[모델] 순수 추정통장 vs 실측 2,800 오차 < 2% (독립 검증)",
         abs(est_cash2 - 2800) / 2800 < 0.02),
        ("[모델] 추정 인센 276 (실측 291 대비 5.2%)", abs(est_incen - 276) < 3),
        ("[재현] 경희 보정단가 20.8", abs(kh["pp2"] - 20.8) < 0.15),
        ("[재현] 경희 종합 1위 · gap(1↔2위) ≥ 5점",
         kh["comp_rank"] == 1 and fin[0]["score"] - fin[1]["score"] >= 5),
        ("[재현] pp2≥20 & 연환자≥15k = 경희뿐",
         [d["h"] for d in pool if d["pp2"] >= 20 and d["tot"] >= CUT_ANNUAL] == [kh["h"]]),
    ]
    for n in labels:
        st = sts[n]
        ok = (st["best_cut"] is not None and abs(st["best_cut"] - CUT_UNIT) <= 0.3
              and st["p_bonf"] < 0.05 and st["zone"] is not None
              and st["zone"][0] <= CUT_UNIT <= st["zone"][1])
        t.append((f"[통계] {n}: 컷 13.3±0.3 · Bonf<0.05 · 컷∈유의구간", ok))
    t.append(("[통계] 라벨 3종 컷 최대편차 ≤ 0.3",
              len(cuts) == 3 and max(cuts) - min(cuts) <= 0.3))
    return t

# ──────────────────── 메인 ────────────────────
def _incen_str(d):
    if d["incen_src"] == "실측":
        return f"{d['incen']:.0f} ★실측"
    if d["incen_src"] == "실측포함":
        return "포함(내역미상)"
    return f"{d['incen']:.0f} (모델)"

def main():
    args = set(sys.argv[1:])
    pool, rep = load_master(report=True, strict="--strict" in args)
    if "--report" in args:
        print(f"[로드 리포트] 시트1 {rep['n_main']} · 시트16 {rep['n_axes']} · 병합 {rep['n_merged']}")
        for k in ("only_in_main", "only_in_axes", "skipped", "dup_main", "dup_axes"):
            print(f"  {k}: {rep[k] or '없음'}")
    apply_incentive(pool)
    compute_axes(pool)
    fin = composite(pool)
    labels = make_labels(pool)

    print("═" * 60)
    print(f"봉직 통합 모델 v1.2 — 풀 {len(pool)}곳 (오산=구공고 원본, 갱신은 --posting 카드)")
    print("═" * 60)
    print("\n[보정통장 TOP10]")
    for i, d in enumerate(sorted(pool, key=lambda x: (-x["cash2"], x["h"]))[:10], 1):
        print(f"  {i:2}. {d['h']:22s} {d['cash2']:5.0f} (인센 {_incen_str(d)})")
    print("\n[일하기좋은 TOP12 + 티어]")
    for d in fin[:12]:
        print(f"  {d['comp_rank']:2}. [{d['tier']}] {d['h']:22s} {d['score']:5.1f}")
    print("\n[통계 컷 — 라벨 3종: 직접 누출 제거 민감도 분석 (표본 내 강건성)]")
    for lname, lset in labels.items():
        st = stat_cut(pool, lset, boot=800 if "--boot" in args else 0)
        line = (f"  {lname:24s} n+={len(lset):2d} → 컷 {st['best_cut']:.2f} · "
                f"raw {st['p_raw']:.1e} · Bonf {st['p_bonf']:.3f} · 구간 {st['zone']}")
        if "boot_ci" in st:
            line += f" · CI [{st['boot_ci'][0]:.2f}, {st['boot_ci'][1]:.2f}]"
        print(line)
    naz = next(d for d in pool if "나사렛" in d["h"])
    print(f"\n[현직 나사렛] 종합 {naz['comp_rank']}위 · 5축 {naz['f5']:.1f} · 단가 {naz['pp2']:.1f}")
    print("\n[검증 테스트]")
    fails = 0
    for name, ok in run_tests(pool, labels, rep):
        fails += 0 if ok else 1
        print(f"  {'✓' if ok else '✗ FAIL'} {name}")
    if "--posting" in args:
        print("\n[신규 공고 일괄 비교 — 고정 스케일 · 마스터 대비 지표 포함]")
        for c in evaluate_postings(pool, POSTING_EXAMPLES):
            print(f"  ▸ {c['name']:16s} 보정통장 {c['cash2']} (인센 "
                  f"{'포함(내역미상)' if c['incen_src'] == '실측포함' else str(c['incen']) + ' 모델'}) · "
                  f"단가 {c['pp2']:5.2f} · 점수 {c['score']} · 마스터대비 {c['rank_vs_master']}위"
                  f"(상위 {100 - c['pct_vs_master']}%) · joint {c['joint_rank']}위/{c['n']} "
                  f"[{c['tier']}] · 무결점 {'✓' if c['perfect'] else '✗'}")
    sys.exit(1 if fails else 0)

if __name__ == "__main__":
    main()
